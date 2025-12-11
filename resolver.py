#UEFI 安全启动数据库解析器。
#解析 PK/KEK/DB/DBX 签名列表并导出证书信息。

import struct
import uuid
from cryptography import x509
from cryptography.hazmat.backends import default_backend
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def parse_signature_list(data):
    offset = 0
    results = []

    while offset < len(data):

        # 至少要有 EFI_SIGNATURE_LIST 的固定头（28 bytes）
        if len(data) - offset < 28:
            break

        SignatureType = uuid.UUID(bytes_le=data[offset:offset+16])
        SignatureListSize, SignatureHeaderSize, SignatureSize = struct.unpack(
            "<III", data[offset+16:offset+28]
        )

        # 简单边界判断（部分主板可能产生损坏结构）
        if SignatureListSize < 28 or offset + SignatureListSize > len(data):
            break

        header_start = offset + 28
        header_end = header_start + SignatureHeaderSize
        header = data[header_start:header_end]

        entry_offset = header_end
        end_of_list = offset + SignatureListSize

        # 遍历结构内每个 SignatureData 条目
        while entry_offset + SignatureSize <= end_of_list:
            entry = data[entry_offset:entry_offset + SignatureSize]

            owner_guid = uuid.UUID(bytes_le=entry[:16])
            signature_data = entry[16:]

            info = {
                "SignatureType": str(SignatureType),
                "OwnerGUID": str(owner_guid),
                "is_x509": False,
                "subject": "",
                "issuer": "",
                "serial_number": "",
                "not_before": "",
                "not_after": "",
                "data_length": len(signature_data),
            }

            # ----------- 尝试解析 X.509（无弃用警告） -------------
            try:
                cert = x509.load_der_x509_certificate(signature_data, default_backend())
                info["is_x509"] = True
                info["subject"] = cert.subject.rfc4514_string()
                info["issuer"] = cert.issuer.rfc4514_string()
                info["serial_number"] = str(cert.serial_number)

                # 使用无警告的新属性（包含 UTC 时区）
                info["not_before"] = cert.not_valid_before_utc.isoformat()
                info["not_after"] = cert.not_valid_after_utc.isoformat()

            except Exception:
                pass  # 非证书，例如 dbx SHA256 条目

            results.append(info)
            entry_offset += SignatureSize

        offset += SignatureListSize

    return results


# ----------- 读取 PK / KEK / DB / DBX 文件 -------------
files = {
    "PK": "PK.bin",
    "KEK": "KEK.bin",
    "db": "db.bin",
    "dbx": "dbx.bin"
}

all_entries = []

for store_name, filename in files.items():
    try:
        with open(filename, "rb") as f:
            data = f.read()
        entries = parse_signature_list(data)
        for e in entries:
            e["Store"] = store_name
        all_entries.extend(entries)
    except FileNotFoundError:
        print(f"{filename} 文件不存在，跳过。")


# ----------- 生成 Excel（自动列宽，无警告） -------------
wb = Workbook()
ws = wb.active
ws.title = "UEFI SecureBoot Certificates"

headers = [
    "Store", "SignatureType", "OwnerGUID", "is_x509",
    "subject", "issuer", "serial_number",
    "not_before", "not_after", "data_length"
]
ws.append(headers)

for e in all_entries:
    ws.append([
        e["Store"],
        e["SignatureType"],
        e["OwnerGUID"],
        e["is_x509"],
        e["subject"],
        e["issuer"],
        e["serial_number"],
        e["not_before"],
        e["not_after"],
        e["data_length"]
    ])

# 自动调列宽
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value is not None:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

output_file = "UEFI_SecureBoot_entries.xlsx"
wb.save(output_file)

print(f"Excel 文件已生成：{output_file}")
